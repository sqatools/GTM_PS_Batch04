import openpyxl
"""
pip install openpyxl
"""

def read_excel_file(filename, n1, n2):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb['Sheet1']
    #print(sheet.__dict__)
    # get specific cell
    cell = sheet.cell(row=n1, column=n2)

    print(cell.value)


#read_excel_file("test_data.xlsx", n1=1, n2=1)

# for i in range(1, 5):
#     read_excel_file("test_data.xlsx", n1=i, n2=1)

def read_excel_file_with_cellname(filename):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb['Sheet1']
    # get specific cell
    cell = sheet['A1']

    print(cell.value)


#read_excel_file_with_cellname("test_data.xlsx")


def read_excel_file_with_maxrow(filename, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    rows = sheet.rows
    for cell in rows:
        print(cell[0].value)


#read_excel_file_with_maxrow("test_data.xlsx", 'Sheet1')


def read_excel_file_with_maxcol(filename, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    max_colums = sheet.columns
    for cell in max_colums:
        print(cell[0].value)


# read_excel_file_with_maxcol("test_data.xlsx", 'Sheet1')
# read_excel_file_with_maxcol("test_data.xlsx", 'Sheet2')


def read_excel_data_with_cell_name(filename, cell_name, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    cell = sheet[cell_name]
    return cell.value

#data = read_excel_data_with_cell_name("test_data.xlsx", "A1", "Sheet2")

#print(data)


#######################
# update the exel sheet data

def update_excel_sheet_data(filename, sheet_name, cell_name, new_data):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    sheet[cell_name] = new_data
    wb.save(filename)

#update_excel_sheet_data("test_data.xlsx", "Sheet1", "A1", "Chennai")
update_excel_sheet_data("test_data.xlsx", "Sheet1", "B12", "=SUM(B2:B10)")

# =SUM(B2:B10)


#Q1 write a python program to enter all even number from 1 to 20 in A colom
#Q2 write a python program to enter a table of 5 in excel sheet
"""
num = 5
for i in range(1, 11):
    v1 = f"{num}*{i}"
    v2 = num*i
    update_excel_sheet_data("test_data.xlsx", "Sheet3", f"A{i}", v1)
    update_excel_sheet_data("test_data.xlsx", "Sheet3", f"B{i}", v2)
"""

#Q1 write a python program to enter all even number from 1 to 20 in A column
count = 1
for i in range(1, 21): # i = 1, 2, 3, 4 count = 1, 2, 3
    if i%2 == 0:
        update_excel_sheet_data("test_data.xlsx", "Sheet3", f"C{count}", i)
        count = count + 1
