import openpyxl
"""
pip install openpyxl
"""
def read_excel_file(filename, n1, n2):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb['Sheet1']
    #print(sheet.__dict__)
    # get specific cell
    cell = sheet.cell(row=n1, column=n2)

    print(cell.value)
#read_excel_file("test_data1.xlsx", n1=1, n2=1)

#for i in range(1, 5):
   # read_excel_file("test_data.xlsx", n1=i, n2=1)

def read_excel_file_with_cellname(filename):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb['Sheet1']
    # get specific cell
    cell = sheet['A1']

    print(cell.value)


#read_excel_file_with_cellname("test_data1.xlsx")

def read_excel_file_with_maxrow(filename, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    rows = sheet.rows
    for cell in rows:
        print(cell[0].value)


#read_excel_file_with_maxrow("test_data1.xlsx", 'Sheet1')

def read_excel_file_with_maxcol(filename, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    max_colums = sheet.columns
    for cell in max_colums:
        print(cell[0].value)


#read_excel_file_with_maxcol("test_data1.xlsx", 'Sheet1')
#read_excel_file_with_maxcol("test_data1.xlsx", 'Sheet2')

def read_excel_data_with_cell_name(filename, cell_name, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    cell = sheet[cell_name]
    return cell.value

#data = read_excel_data_with_cell_name("test_data1.xlsx", "A1", "Sheet2")

#print(data)

def update_excel_sheet_data(filename, sheet_name, cell_name, new_data):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    sheet[cell_name] = new_data
    wb.save(filename)

update_excel_sheet_data("test_data1.xlsx", "Sheet1", "A1", "Chennai")
#update_excel_sheet_data("test_data1.xlsx", "Sheet1", "B12", "=SUM(B2:B10)")

 #=SUM(B2:B10)

 #Q1 write a python program to enter all even number from 1 to 20 in A colomn
#Q2 write a python program to enter a table of 5 in excel sheet

#Q1 write a python program to enter all even number from 1 to 20 in A colom

def even_number(filename,sheet_name,cell_name,new_data):
  wb = openpyxl.load_workbook(filename)
  sheet = wb[sheet_name]
  sheet[cell_name] = new_data
  wb.save(filename)
for cell in rows:

even_number(("test_data1.xlsx", "Sheet1", "c2", "Chennai"))

#
wb = openpyxl.Workbook()
sheet = wb['Sheet'] # or wb.active
n = int(input('enter n| '))
