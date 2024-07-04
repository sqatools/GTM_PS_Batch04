
"""
pip install openpyxl
"""
import openpyxl


#read with row and column

# import openpyxl
# def read_excel_file(filename,row,column):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     cell =sheet.cell(row=row,column=column)
#     print(cell.value)

#read_excel_file('test_data.xlsx',row=1,column=6)

# for i in range(1, 6):
#      read_excel_file("test_data.xlsx", row=i, column=1)
#
#####################################################################################

#Read with cellname

# import openpyxl
# def read_with_cell_name(filename):
#
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     cell = sheet['I1']
#     print(cell.value)
#
# read_with_cell_name('test_data.xlsx')

#########################################################################

#read excel file with max row

# import openpyxl
#
# def read_excel_with_max_rows(filename):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     row = sheet.rows
#     for cell in row:
#         print(cell[1].value)
# read_excel_with_max_rows('test_data.xlsx')

##########################################################################

#read excel file with max column:
#
# import openpyxl
#
# def read_excel_file_max_cols(filename):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     column = sheet.columns
#     for cell in column:
#         print(cell[2].value)
# read_excel_file_max_cols('test_data.xlsx')

###########################################################################################
#read_excel_data_with_cell_name

# def read_excel_data_with_cell_name(filename, cell_name, sheet_name):
#     # loaded excel workbook
#     wb = openpyxl.load_workbook(filename)
#     # open sheet number
#     sheet = wb[sheet_name]
#     # get specific cell
#     cell = sheet[cell_name]
#     return cell.value
#
# #data = read_excel_data_with_cell_name("test_data.xlsx", "A1", "Sheet2")
#
# #print(data)
#############################################################################################


# # update the exel sheet data

# import openpyxl
#
# def update_excel_sheet(filename ,cellname,new_data):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     sheet[cellname] = new_data
#     wb.save(filename)
#
# update_excel_sheet("test_data.xlsx" , "F2", "varkala")

#############################EXERCISE###################################

# #Q1 write a python program to enter all even number from 1 to 20 in A colom

# #Q2 write a python program to enter a table of 5 in excel sheet

#Q1 write a python program to enter all even number from 1 to 20 in A colom
#Q2 write a python program to enter a table of 5 in excel sheet

# import openpyxl
#
# def update_excel_sheet_data(filename, sheet_name, cell_name, new_data):
#     wb = openpyxl.load_workbook(filename)
#     sheet = wb[sheet_name]
#     sheet[cell_name] = new_data
#     wb.save(filename)
# num = 5
# for i in range(1, 11):
#     v1 = f"{num}*{i}"
#     v2 = num*i
#     update_excel_sheet_data("test_data2.xlsx", "Sheet1", f"A{i}", v1)
#     update_excel_sheet_data("test_data2.xlsx", "Sheet1", f"B{i}", v2)



# #Q1 write a python program to enter all even number from 1 to 20 in A column

import openpyxl

def update_excel_sheet_data(filename, sheet_name, cell_name, new_data):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    sheet[cell_name] = new_data
    wb.save(filename)
count = 1
for i in range(1, 21): # i = 1, 2, 3, 4 count = 1, 2, 3
    if i%2 == 0:
        update_excel_sheet_data("test_data2.xlsx", "Sheet1", f"C{count}", i)
        count = count + 1
