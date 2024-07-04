import random

import openpyxl

def read_excwl_sheet(filename,n1,n2):
    wb=openpyxl.load_workbook(filename)
    sheet=wb["Sheet1"]
    cell=sheet.cell(row=n1,column=n2)
    print(cell.value)

#read_excwl_sheet("xl_test.xlsx",n1=5,n2=2)

#for i in range(1,10):
    #read_excwl_sheet("xl_test.xlsx",n1=i,n2=2)

def read_excel_file(filename):
    wb=openpyxl.load_workbook(filename)
    sheet=wb["Sheet1"]
    cell=sheet["B4"]
    print(cell.value)
#read_excel_file("xl_test.xlsx")


def read_excel_maxrow(filename,sheet_name):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheet_name]
    rows=sheet.rows
    for cell in rows:
        print("city: ",cell[0].value,"pin:",cell[1].value)
#read_excel_maxrow("xl_test.xlsx","Sheet1")

def read_excel_file_maxcolumns(filename,sheet_name):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheet_name]
    columns=sheet.columns
    for cell in columns:
        print(cell[0].value)

#read_excel_file_maxcolumns("xl_test.xlsx","Sheet1")


def read_excel_file_with_cellname(filename,cell_name,sheet_name):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheet_name]
    cell=sheet[cell_name]
    return cell.value

#data=read_excel_file_with_cellname("xl_test.xlsx","B3","Sheet1")
#print(data)

def read_excel_sheet_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)
    sheet=wb[sheetname]
    for i in range(2,21,2):
        sheet["A"+ str(i//2)]=i

    wb.save(filename)

read_excel_sheet_data("xl_test.xlsx","Sheet1")

def read_excel_file_with_maxrow(filename, sheet_name):
    # loaded excel workbook
    wb = openpyxl.load_workbook(filename)
    # open sheet number
    sheet = wb[sheet_name]
    # get specific cell
    rows = sheet.rows
    for cell in range(1,21):
        if cell%2==0:
            wb.save(cell.value)

#read_excel_file_with_maxrow("xl_test.xlsx","Sheet1")
#read_excel_file_with_maxrow("xl_test.xlsx","Sheet1")


# import openpyxl
#
# # Create a new Workbook
# wb = openpyxl.Workbook()
#
# # Select the active sheet
# sheet = wb.active
# sheet.title = 'Even Numbers'
#
# # Iterate through even numbers from 1 to 20 and write them to column A
# for num in range(2, 21, 2):  # range(start, stop, step)
#     sheet['A' + str(num // 2)] = num  # num // 2 gives row number
#
# # Save the workbook
# wb.save('even_numbers.xlsx')
#
# print("Excel file 'even_numbers.xlsx' has been created with even numbers from 1 to 20.")
